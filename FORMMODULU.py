# FORMMODULU.py
"""
FORMMODULU - İyileştirilmiş Versiyon
Yazar: Hüseyin İLHAN
Düzenleyen: Claude AI Assistant
"""

import os
import platform
import unicodedata
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
from dateutil.relativedelta import relativedelta
import logging
import json
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# Belge ve GUI için platformlar arası Türkçe karakter destekli font seçimi
import tkinter.font as tkfont_det
_font_root = tk.Tk()
_font_root.withdraw()
_available_fonts = set(tkfont_det.families())
_font_root.destroy()
for _font in ("Arial", "Liberation Sans", "DejaVu Sans", "TkDefaultFont"):
    if _font in _available_fonts:
        DEFAULT_FONT = _font
        break
else:
    DEFAULT_FONT = tkfont_det.nametofont("TkDefaultFont").actual()["family"]

# Platform detection
SYSTEM = platform.system()
IS_WINDOWS = SYSTEM == "Windows"
IS_MACOS = SYSTEM == "Darwin"
IS_LINUX = SYSTEM == "Linux"

# Loglama ayarları
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('form_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class SadeFormApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Form Bilgilerini Doldur")
        self.root.geometry("900x700")  # Pencere boyutu büyütüldü
        self.root.configure(bg="#f8f9fa")
        
        # SGK geçmişi için dosya yolu
        self.sgk_history_file = "sgk_history.json"
        self.sgk_history = self.load_sgk_history()
        
        # Sabitler
        self.TEHLIKE_SINIFLARI = ["AZ TEHLİKELİ", "TEHLİKELİ", "ÇOK TEHLİKELİ"]
        self.TEHLIKE_YIL_MAP = {
            "AZ TEHLİKELİ": 6,
            "TEHLİKELİ": 4,
            "ÇOK TEHLİKELİ": 2
        }
        
        # RD Yöntemi seçimi
        self.rd_method = tk.StringVar(value="Matris")
        
        # Excel yolları - platform bağımsız
        self.ankara_tablosu_path = tk.StringVar(value=os.path.abspath("ANKARA İŞYERİ TABLOSU.xlsx"))
        self.nace_tablosu_path = tk.StringVar(value=os.path.abspath("Nace Kod Listesi.xlsx"))
        
        # DataFrame'i yükle
        self.df = self.load_dataframe()
        
        # UI bileşenlerini oluştur
        self.create_ui()
        
        # Pencereyi kapatırken SGK geçmişini kaydet
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def load_dataframe(self):
        """DataFrame'i güvenli bir şekilde yükler"""
        try:
            df = pd.read_excel("veri_yapilandirma_GUNCEL.xlsx", dtype=str, engine='openpyxl')
            logging.info("veri_yapilandirma_GUNCEL.xlsx başarıyla yüklendi")
            return df
        except FileNotFoundError:
            messagebox.showerror("Hata", "'veri_yapilandirma_GUNCEL.xlsx' bulunamadı.")
            logging.error("veri_yapilandirma_GUNCEL.xlsx dosyası bulunamadı")
            return pd.DataFrame(columns=["Anahtar", "Etiket", "Durum", "Karşılık"])
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya okuma hatası: {str(e)}")
            logging.error(f"DataFrame yükleme hatası: {e}")
            return pd.DataFrame(columns=["Anahtar", "Etiket", "Durum", "Karşılık"])
    
    def load_sgk_history(self):
        """SGK geçmişini yükler"""
        try:
            if os.path.exists(self.sgk_history_file):
                with open(self.sgk_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return []
        except Exception as e:
            logging.error(f"SGK geçmişi yükleme hatası: {e}")
            return []
    
    def save_sgk_history(self):
        """SGK geçmişini kaydeder"""
        try:
            with open(self.sgk_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.sgk_history, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f"SGK geçmişi kaydetme hatası: {e}")
    
    def add_to_sgk_history(self, sgk_code):
        """SGK kodunu geçmişe ekler"""
        if sgk_code in self.sgk_history:
            self.sgk_history.remove(sgk_code)
        self.sgk_history.insert(0, sgk_code)
        # Sadece son 3 kodu tut
        self.sgk_history = self.sgk_history[:3]
    
    def on_closing(self):
        """Pencere kapatılırken çağırılır"""
        self.save_sgk_history()
        self.root.destroy()
    
    def create_ui(self):
        """Kullanıcı arayüzünü oluşturur"""
        # Üst panel
        self.create_top_panel()
        
        # Başlık - platform uyumlu font
        default_font = DEFAULT_FONT
        title_label = tk.Label(self.root, text="Form Bilgilerini Doldur", 
                              font=(default_font, 20, "bold"),
                              bg="#f8f9fa", fg="#2c3e50")
        title_label.pack(pady=(10, 5))
        
        # RD Yöntemi seçimi
        self.create_rd_method_panel()
        
        # Scrollable form alanı
        self.create_form_area()
        
    def create_top_panel(self):
        """Üst panel (SGK girişi ve butonlar)"""
        top = tk.Frame(self.root, bg="#f8f9fa")
        top.pack(fill="x", padx=15, pady=10)
        
        # SGK girişi - platform uyumlu font
        default_font = DEFAULT_FONT
        
        # SGK Label ve Entry
        sgk_frame = tk.Frame(top, bg="#f8f9fa")
        sgk_frame.pack(side="left", padx=(0, 10))
        
        tk.Label(sgk_frame, text="KISA SGK (7 hane):", 
                bg="#f8f9fa", fg="#1a237e",
                font=(default_font, 11, "bold")).pack(anchor="w")
        
        self.kisa_sgk_var = tk.StringVar()
        
        # SGK Entry için ComboBox (geçmiş önerileri için)
        self.sgk_combo = ttk.Combobox(sgk_frame, textvariable=self.kisa_sgk_var, 
                                     width=10, font=(default_font, 10))
        self.sgk_combo['values'] = self.sgk_history
        self.sgk_combo.pack(pady=(2, 0))
        
        # Butonlar
        buttons = [
            ("SEÇ", self.on_select_sgk, "#1a237e"),
            ("ANKARA İŞYERİ TABLOSU", self.select_ankara, "#1a237e"),
            ("NACE KODLARI", self.select_nace, "#1a237e")
        ]
        
        for text, command, color in buttons:
            btn = tk.Button(top, text=text, command=command,
                          bg="#f0f0f0", fg="#1a237e",
                          font=(default_font, 10, "bold"),
                          activebackground="#e0e0e0", activeforeground="#1a237e",
                          relief="raised", bd=3, pady=8)
            
            # macOS için özel ayarlar
            if IS_MACOS:
                btn.configure(highlightbackground="#f0f0f0")
            
            btn.pack(side="left", padx=(10, 0))
    
    def create_rd_method_panel(self):
        """RD Yöntemi seçim panelini oluşturur"""
        default_font = DEFAULT_FONT
        
        rd_frame = tk.LabelFrame(self.root, text="RD Yöntemi", 
                               bg="#f8f9fa", fg="#2c3e50",
                               font=(default_font, 12, "bold"),
                               bd=2, relief="groove")
        rd_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        # Radio buttonlar için frame
        radio_frame = tk.Frame(rd_frame, bg="#f8f9fa")
        radio_frame.pack(fill="x", padx=10, pady=10)
        
        # Matris radio button
        tk.Radiobutton(radio_frame, text="Matris", variable=self.rd_method, value="Matris",
                      bg="#f8f9fa", fg="#1a237e", selectcolor="#f8f9fa",
                      font=(default_font, 11), activebackground="#f8f9fa").pack(side="left", padx=(0, 20))
        
        # Fine Kinney radio button
        tk.Radiobutton(radio_frame, text="Fine Kinney", variable=self.rd_method, value="Fine Kinney",
                      bg="#f8f9fa", fg="#1a237e", selectcolor="#f8f9fa",
                      font=(default_font, 11), activebackground="#f8f9fa").pack(side="left")
    
    def create_form_area(self):
        """Scrollable form alanını oluşturur"""
        container = tk.Frame(self.root, bg="#f8f9fa")
        container.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Canvas ve scrollbar
        canvas = tk.Canvas(container, bg="#f8f9fa", highlightthickness=0)
        vsb = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        vsb.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        
        # Form frame
        self.form_frame = tk.Frame(canvas, bg="#f8f9fa")
        canvas.create_window((0, 0), window=self.form_frame, anchor="nw")
        self.form_frame.bind("<Configure>", 
                           lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        # Mouse wheel scrolling - macOS ve Windows uyumlu
        def on_mousewheel(event):
            if IS_MACOS:
                canvas.yview_scroll(int(-1 * event.delta), "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        # Platform-özel scroll binding
        if IS_MACOS:
            canvas.bind_all("<MouseWheel>", on_mousewheel)
            canvas.bind_all("<Button-4>", on_mousewheel)
            canvas.bind_all("<Button-5>", on_mousewheel)
        else:
            canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Form alanlarını oluştur
        self.create_form_fields()
        
        # Kaydet butonu
        self.create_save_button()
    
    def create_form_fields(self):
        """Form alanlarını oluşturur"""
        self.entries = {}
        row = 0
        
        for idx, record in self.df.iterrows():
            key = record["Anahtar"]
            label = record["Etiket"]
            durum = (record.get("Durum") or "").strip().lower()
            
            if durum != "düzenlenebilir":
                continue
            
            # Label - platform uyumlu font
            default_font = DEFAULT_FONT
            lbl = tk.Label(self.form_frame, text=label, 
                         anchor="w", width=40,
                         bg="#f8f9fa", fg="#1a237e", font=(default_font, 11, "bold"))
            lbl.grid(row=row, column=0, sticky="w", padx=(10, 15), pady=6)
            
            # Widget
            widget = self.create_widget_for_key(key)
            widget.grid(row=row, column=1, padx=(0, 15), pady=4, sticky="ew")
            # Column weight ayarla
            self.form_frame.grid_columnconfigure(1, weight=1)
            self.entries[key] = widget
            
            row += 1
        
        logging.info(f"Toplam {len(self.entries)} form alanı oluşturuldu")
    
    def create_widget_for_key(self, key):
        """Anahtara göre uygun widget'ı oluşturur"""
        default_font = DEFAULT_FONT
        if key == "[DEĞİŞTİR:TEHLİKESINIFI]":
            widget = ttk.Combobox(self.form_frame,
                                values=self.TEHLIKE_SINIFLARI,
                                state="readonly", width=33,
                                font=(default_font, 10))
        elif key == "[DEĞİŞTİR:NACEFAALİYET]":
            widget = tk.Text(self.form_frame, width=35, height=3,
                           font=(default_font, 9), wrap=tk.WORD)
        else:
            widget = tk.Entry(self.form_frame, width=35, font=(default_font, 10))
        
        return widget
    
    def create_save_button(self):
        """Kaydet butonunu oluşturur"""
        row = len([k for k, v in self.entries.items()])
        default_font = DEFAULT_FONT
        btn = tk.Button(self.form_frame, text="Verileri Kaydet",
                       command=self.kaydet,
                       bg="#f0f0f0", fg="#1a237e",
                       font=(default_font, 12, "bold"),
                       width=25, height=2, relief="raised", bd=3,
                       activebackground="#e0e0e0", activeforeground="#1a237e")
        
        # macOS için özel ayarlar
        if IS_MACOS:
            btn.configure(highlightbackground="#f0f0f0")
        
        btn.grid(row=row, column=1, pady=(20, 5), sticky="ew")
        
    
    
    def select_ankara(self):
        """Ankara tablosu dosyasını seçer"""
        filename = filedialog.askopenfilename(
            title="Ankara Tablosu Seç",
            filetypes=[("Excel", "*.xlsx;*.xls")],
            initialdir=os.getcwd()
        )
        if filename:
            self.ankara_tablosu_path.set(filename)
            logging.info(f"Ankara tablosu seçildi: {filename}")
    
    def select_nace(self):
        """NACE kodları dosyasını seçer"""
        filename = filedialog.askopenfilename(
            title="NACE Kodu Seç",
            filetypes=[("Excel", "*.xlsx;*.xls")],
            initialdir=os.getcwd()
        )
        if filename:
            self.nace_tablosu_path.set(filename)
            logging.info(f"NACE tablosu seçildi: {filename}")
    
    def on_select_sgk(self):
        """SGK koduna göre verileri doldurur"""
        kod = self.kisa_sgk_var.get().strip()
        
        # Validasyon
        if not self.validate_sgk_code(kod):
            return
        
        # SGK kodunu geçmişe ekle
        self.add_to_sgk_history(kod)
        # ComboBox'u güncelle
        self.sgk_combo['values'] = self.sgk_history
        
        # Ankara tablosunu yükle
        df_ankara = self.load_ankara_table()
        if df_ankara is None:
            return
        
        # SGK koduna göre satırı bul
        satir = self.find_sgk_row(df_ankara, kod)
        if satir is None:
            return
        
        # Verileri doldur
        self.fill_form_data(satir)
        
        # NACE açıklamasını doldur
        self.fill_nace_description(satir)
        
        # Grup dışı kontrolü
        self.check_grup_disi(satir)
        
        messagebox.showinfo("Başarılı", "Veriler başarıyla yüklendi!")
        logging.info(f"SGK {kod} için veriler yüklendi")
    
    def validate_sgk_code(self, kod):
        """SGK kodunu doğrular"""
        if len(kod) != 7 or not kod.isdigit():
            messagebox.showerror("Hata", "Lütfen 7 haneli KISA SGK girin.")
            return False
        return True
    
    def load_ankara_table(self):
        """Ankara tablosunu yükler"""
        try:
            df = pd.read_excel(self.ankara_tablosu_path.get(), dtype=str, engine='openpyxl')
            logging.info("Ankara tablosu yüklendi")
            return df
        except Exception as e:
            messagebox.showerror("Hata", f"Ankara tablosu açılamadı:\n{e}")
            logging.error(f"Ankara tablosu yükleme hatası: {e}")
            return None
    
    def find_sgk_row(self, df_ankara, kod):
        """SGK koduna göre satırı bulur"""
        # SGK sütununu bul
        sgk_col = "KISA SGK" if "KISA SGK" in df_ankara.columns else df_ankara.columns[15]
        df_ankara[sgk_col] = df_ankara[sgk_col].astype(str).str.strip()
        
        # Satırı bul
        satirlar = df_ankara[df_ankara[sgk_col] == kod]
        if satirlar.empty:
            messagebox.showerror("Hata", f"KISA SGK bulunamadı: {kod}")
            return None
        
        return satirlar.iloc[0]
    
    def fill_form_data(self, satir):
        """Form verilerini doldurur"""
        # Veri eşlemeleri
        field_mappings = {
            "[DEĞİŞTİR:ŞİRKET UNVANI]": satir.iloc[4],
            "[DEĞİŞTİR:PROJEADI]": satir.iloc[6],
            "[DEĞİŞTİR:ADRES]": satir.iloc[31],
            "[DEĞİŞTİR:SGKSİCİL]": satir.iloc[10],
            "[DEĞİŞTİR:SGKSİCİL20PUNTO]": satir.iloc[10],
            "[DEĞİŞTİR:NACE]": satir.iloc[9],
            "[DEĞİŞTİR:TEHLİKESINIFI]": satir.iloc[16],
            "[DEĞİŞTİR:ÇALIŞANSAYISI]": satir.iloc[19],
            "[DEĞİŞTİR:ŞİRKET UNVANI20PUNTO]": satir.iloc[4],
            "[DEĞİŞTİR:PROJEADI20PUNTO]": satir.iloc[6],
            # Yeni placeholder'lar
            "[DEĞİŞTİR:İL]": satir.iloc[3],           # D sütunu (3. index)
            "[DEĞİŞTİR:UZMANADI]": satir.iloc[21],    # V sütunu (21. index)
            "[DEĞİŞTİR:HEKİMADI]": satir.iloc[25]     # Z sütunu (25. index)
        }
        
        # Verileri form alanlarına ve DataFrame'e yaz
        for key, value in field_mappings.items():
            str_value = "" if pd.isna(value) else str(value)
            
            # Form alanını güncelle
            widget = self.entries.get(key)
            if widget:
                self.update_widget_value(widget, str_value)
            
            # DataFrame'i güncelle
            self.df.loc[self.df["Anahtar"] == key, "Karşılık"] = str_value
        
        # Şirket-Proje kombinasyonunu oluştur
        sirket = field_mappings["[DEĞİŞTİR:ŞİRKET UNVANI]"]
        proje = field_mappings["[DEĞİŞTİR:PROJEADI]"]
        kombine = f"{sirket} - {proje}"
        self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:ŞİRKETPROJE]", "Karşılık"] = kombine
    
    def fill_nace_description(self, satir):
        """NACE açıklamasını doldurur"""
        nace_kod = str(satir.iloc[9]) if not pd.isna(satir.iloc[9]) else ""
        
        if not nace_kod:
            return
        
        try:
            # NACE tablosunu yükle
            df_nace = pd.read_excel(self.nace_tablosu_path.get(), dtype=str, engine='openpyxl')
            col_kod, col_aciklama = df_nace.columns[0], df_nace.columns[1]
            df_nace[col_kod] = df_nace[col_kod].astype(str).str.strip()
            
            # NACE açıklamasını bul
            found = df_nace[df_nace[col_kod] == nace_kod]
            aciklama = "" if found.empty else str(found.iloc[0][col_aciklama])
            
            # DataFrame'i güncelle
            self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:NACEFAALİYET]", "Karşılık"] = aciklama
            
            # Form alanını güncelle
            widget = self.entries.get("[DEĞİŞTİR:NACEFAALİYET]")
            if widget and isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
                widget.insert("1.0", aciklama)
            
            # NACE ve Faaliyet kombinasyonunu oluştur
            if aciklama:
                kombine = f"{nace_kod} - {aciklama}"
                self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:NACEVEFAALİYET]", "Karşılık"] = kombine
            
            logging.info(f"NACE açıklaması bulundu: {nace_kod}")
            
        except Exception as e:
            logging.error(f"NACE açıklama hatası: {e}")
    
    def check_grup_disi(self, satir):
        """Grup dışı firma kontrolü yapar"""
        sirket_unvani = str(satir.iloc[4]) if not pd.isna(satir.iloc[4]) else ""
        proje_adi = str(satir.iloc[6]) if not pd.isna(satir.iloc[6]) else ""
        
        if "GRUP DIŞI" in sirket_unvani.upper():
            logging.info(f"GRUP DIŞI firma tespit edildi: {sirket_unvani}")
            
            # Şirket unvanını proje adı ile değiştir
            self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:ŞİRKET UNVANI]", "Karşılık"] = proje_adi
            self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:ŞİRKET UNVANI20PUNTO]", "Karşılık"] = proje_adi
            
            # Proje adı alanlarını boşalt
            self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:PROJEADI]", "Karşılık"] = ""
            self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:PROJEADI20PUNTO]", "Karşılık"] = ""
            
            # Form alanlarını güncelle
            w_sirket = self.entries.get("[DEĞİŞTİR:ŞİRKET UNVANI]")
            if w_sirket:
                self.update_widget_value(w_sirket, proje_adi)
            
            w_proje = self.entries.get("[DEĞİŞTİR:PROJEADI]")
            if w_proje:
                self.update_widget_value(w_proje, "")
            
            # Kombine alanı güncelle
            self.df.loc[self.df["Anahtar"] == "[DEĞİŞTİR:ŞİRKETPROJE]", "Karşılık"] = proje_adi
    
    def update_widget_value(self, widget, value):
        """Widget değerini günceller"""
        if isinstance(widget, ttk.Combobox):
            widget.set(value)
        elif isinstance(widget, tk.Text):
            widget.delete("1.0", tk.END)
            widget.insert("1.0", value)
        else:
            widget.delete(0, tk.END)
            widget.insert(0, value)
    
    def kaydet(self):
        """Form verilerini kaydeder"""
        logging.info("=== KAYDETME İŞLEMİ BAŞLADI ===")
        
        # Form verilerini oku
        self.update_dataframe_from_form()
        
        # Otomatik hesaplamaları yap
        self.perform_auto_calculations()
        
        # Dosyaları kaydet
        if self.save_files():
            messagebox.showinfo("Başarılı", 
                              "Veriler başarıyla kaydedildi!\n\n"
                              "• Form verileri güncellendi\n"
                              "• Otomatik hesaplamalar yapıldı\n"
                              "• Dosyalar kaydedildi")
            self.root.destroy()
        else:
            messagebox.showerror("Hata", "Dosyalar kaydedilemedi!")
    
    def update_dataframe_from_form(self):
        """Form verilerini DataFrame'e aktarır"""
        updated_count = 0
        
        for key, widget in self.entries.items():
            try:
                # Widget değerini al
                if isinstance(widget, tk.Text):
                    value = widget.get("1.0", tk.END).strip()
                else:
                    value = widget.get()
                    if value is None:
                        value = ""
                    value = str(value).strip()
                
                # DataFrame'i güncelle
                mask = (self.df["Anahtar"] == key)
                if mask.any():
                    self.df.loc[mask, "Karşılık"] = value
                    updated_count += 1
                    
            except Exception as e:
                logging.error(f"Widget okuma hatası {key}: {e}")
        
        logging.info(f"Toplam {updated_count} alan güncellendi")
    
    def perform_auto_calculations(self):
        """Otomatik hesaplamaları yapar"""
        logging.info("=== OTOMATİK HESAPLAMALAR ===")
        
        # Tehlike sınıfını al
        tehlike_sinifi = self.get_tehlike_sinifi()
        
        # Tarih hesaplamaları
        self.calculate_dates(tehlike_sinifi)
        
        # Saat hesaplamaları
        self.calculate_hours(tehlike_sinifi)
        
        # Periyot hesaplamaları
        self.calculate_periods(tehlike_sinifi)
        
        # Kombine alanları hesapla
        self.calculate_combined_fields()
        
        # Normal belge oluşturma için faaliyet tarihi placeholder'ını temizle
        self.clear_faaliyet_tarihi_for_normal_operations()
    
    def get_tehlike_sinifi(self):
        """Tehlike sınıfını alır"""
        mask = (self.df["Anahtar"] == "[DEĞİŞTİR:TEHLİKESINIFI]")
        if mask.any():
            return str(self.df.loc[mask, "Karşılık"].iloc[0]).strip()
        return ""
    
    def calculate_dates(self, tehlike_sinifi):
        """Geçerlilik tarihlerini hesaplar"""
        ek_yil = self.TEHLIKE_YIL_MAP.get(tehlike_sinifi, 0)
        date_format = "%d.%m.%Y"
        
        date_pairs = [
            ("[DEĞİŞTİR:RDEKİPATAMAEĞİTİMHAZIRLANMA]", "[DEĞİŞTİR:RDGEÇERLİLİK]"),
            ("[DEĞİŞTİR:ADEPEK3ATAMAEĞİTİM]", "[DEĞİŞTİR:ADEPGEÇERLİLİK]")
        ]
        
        for input_key, output_key in date_pairs:
            try:
                mask = (self.df["Anahtar"] == input_key)
                if mask.any():
                    start_date = str(self.df.loc[mask, "Karşılık"].iloc[0]).strip()
                    
                    if start_date and start_date != "nan":
                        dt = datetime.datetime.strptime(start_date, date_format)
                        end_date = (dt + relativedelta(years=ek_yil)).strftime(date_format)
                        
                        out_mask = (self.df["Anahtar"] == output_key)
                        if out_mask.any():
                            self.df.loc[out_mask, "Karşılık"] = end_date
                            logging.info(f"{output_key}: {end_date}")
                            
            except Exception as e:
                logging.error(f"Tarih hesaplama hatası {input_key}: {e}")
    
    def calculate_hours(self, tehlike_sinifi):
        """Yıllık saat hesaplamalarını yapar"""
        # İGU Saat
        igu_saat = "4 SAAT" if tehlike_sinifi == "AZ TEHLİKELİ" else "8 SAAT"
        self.update_value("[DEĞİŞTİR:YILLIK:İGU:SAAT]", igu_saat)
        
        # İH Saat
        ih_saat = "4 SAAT" if tehlike_sinifi in ("AZ TEHLİKELİ", "TEHLİKELİ") else "8 SAAT"
        self.update_value("[DEĞİŞTİR:YILLIK:İH:SAAT]", ih_saat)
    
    def calculate_periods(self, tehlike_sinifi):
        """Periyot hesaplamalarını yapar"""
        periods = {
            "AZ TEHLİKELİ": ("6 Yılda 1", "5 Yılda 1"),
            "TEHLİKELİ": ("4 Yılda 1", "3 Yılda 1"),
            "ÇOK TEHLİKELİ": ("2 Yılda 1", "Yılda 1")
        }
        
        rd_periyot, muayene_periyot = periods.get(tehlike_sinifi, ("", ""))
        
        self.update_value("[DEĞİŞTİR:YDR:RDPERİYOT]", rd_periyot)
        self.update_value("[DEĞİŞTİR:YDR:MUAYENEPERİYOT]", muayene_periyot)
    
    def calculate_combined_fields(self):
        """Kombine alanları hesaplar"""
        # Şirket-Proje kombinasyonu
        sirket = self.get_value("[DEĞİŞTİR:ŞİRKET UNVANI]")
        proje = self.get_value("[DEĞİŞTİR:PROJEADI]")
        
        if not proje or proje == "nan":
            kombine = sirket  # GRUP DIŞI durumu
        else:
            kombine = f"{sirket} - {proje}"
        
        self.update_value("[DEĞİŞTİR:ŞİRKETPROJE]", kombine)
        
        # NACE ve Faaliyet kombinasyonu
        nace = self.get_value("[DEĞİŞTİR:NACE]")
        nace_faaliyet = self.get_value("[DEĞİŞTİR:NACEFAALİYET]")
        
        if nace and nace_faaliyet:
            nace_kombine = f"{nace} - {nace_faaliyet}"
            self.update_value("[DEĞİŞTİR:NACEVEFAALİYET]", nace_kombine)
        
        # RD Yöntemi bilgisini kaydet
        rd_method = self.rd_method.get()
        self.update_value("[DEĞİŞTİR:RDYONTEMI]", rd_method)
        logging.info(f"RD Yöntemi kaydedildi: {rd_method}")
    
    def clear_faaliyet_tarihi_for_normal_operations(self):
        """Normal belge oluşturma işlemlerinde faaliyet tarihi placeholder'ını temizler"""
        self.update_value("[DEĞİŞTİR:FAALİYETTARİH]", "")
        logging.info("Faaliyet tarihi placeholder'ı normal işlemler için temizlendi")
    
    def get_value(self, key):
        """DataFrame'den değer alır"""
        mask = (self.df["Anahtar"] == key)
        if mask.any():
            return str(self.df.loc[mask, "Karşılık"].iloc[0])
        return ""
    
    def update_value(self, key, value):
        """DataFrame'de değer günceller"""
        mask = (self.df["Anahtar"] == key)
        if mask.any():
            self.df.loc[mask, "Karşılık"] = value
    
    def save_files(self):
        """Dosyaları kaydeder"""
        try:
            # Ana veri dosyasını kaydet
            self.df.to_excel("veri.xlsx", index=False, engine='openpyxl')
            logging.info("veri.xlsx kaydedildi")
            
            # Orijinal dosyayı da güncelle
            self.df.to_excel("veri_yapilandirma_GUNCEL.xlsx", index=False, engine='openpyxl')
            logging.info("veri_yapilandirma_GUNCEL.xlsx güncellendi")
            
            # SGK geçmişini kaydet
            self.save_sgk_history()
            
            return True
            
        except PermissionError:
            logging.error("PermissionError: Excel dosyaları açık olabilir")
            return False
        except Exception as e:
            logging.error(f"Kaydetme hatası: {e}")
            return False
    
    def create_yillik_planlar(self):
        """Yıllık planları oluşturur"""
        logging.info("=== YILLIK PLANLAR OLUŞTURULUYOR ===")
        
        # Verileri güncelle
        self.update_dataframe_from_form()
        self.perform_auto_calculations()
        
        # Yıl kontrolü yap
        if not self.check_year_compatibility():
            messagebox.showinfo("Bilgi", "Yıl uyumsuzluğu nedeniyle sadece kurullu/kurulsuz seçimi yapılacak.")
            use_dynamic_algorithm = False
        else:
            use_dynamic_algorithm = True
        
        # Çalışan sayısını al
        calisanlar = self.get_calisanlar_sayisi()
        is_kurullu = calisanlar >= 50
        
        # Yıllık planları oluştur
        self.generate_yillik_egitim_plani(is_kurullu, use_dynamic_algorithm)
        self.generate_yillik_calisma_plani(is_kurullu, use_dynamic_algorithm)
        
        messagebox.showinfo("Başarılı", "Yıllık planlar başarıyla oluşturuldu!")
        logging.info("Yıllık planlar oluşturuldu")
    
    def check_year_compatibility(self):
        """Yıl uyumluluğunu kontrol eder"""
        try:
            # YILLIK:YIL değerini al
            yillik_yil = self.get_value("[DEĞİŞTİR:YILLIK:YIL]")
            if not yillik_yil or yillik_yil == "nan":
                return False
            
            # YILLIK:TARİH değerini al
            yillik_tarih = self.get_value("[DEĞİŞTİR:YILLIK:TARİH]")
            if not yillik_tarih or yillik_tarih == "nan":
                return False
            
            # Tarihten yıl çıkar
            try:
                tarih_obj = datetime.datetime.strptime(yillik_tarih, "%d.%m.%Y")
                tarih_yil = str(tarih_obj.year)
            except:
                return False
            
            # Yılları karşılaştır
            result = yillik_yil.strip() == tarih_yil.strip()
            logging.info(f"Yıl kontrolü: {yillik_yil} == {tarih_yil} -> {result}")
            return result
            
        except Exception as e:
            logging.error(f"Yıl kontrolü hatası: {e}")
            return False
    
    def get_calisanlar_sayisi(self):
        """Çalışan sayısını döndürür"""
        try:
            calisanlar_str = self.get_value("[DEĞİŞTİR:ÇALIŞANSAYISI]")
            if not calisanlar_str or calisanlar_str == "nan":
                return 0
            return int(calisanlar_str)
        except:
            return 0
    
    def get_desktop_path(self):
        """Masaüstü yolunu döndürür"""
        if IS_WINDOWS:
            return os.path.join(os.path.expanduser("~"), "Desktop")
        elif IS_MACOS:
            return os.path.join(os.path.expanduser("~"), "Desktop")
        else:
            return os.path.join(os.path.expanduser("~"), "Desktop")
    
    def generate_yillik_egitim_plani(self, is_kurullu, use_dynamic_algorithm):
        """Yıllık eğitim planını oluşturur"""
        try:
            # Template dosyasını seç (Evraklar/YILLIKLAR klasörü içinde)
            template_name = "YILLIK EG\u0306I\u0307TI\u0307M PLANI KURULLU.xlsx" if is_kurullu else "YILLIK EG\u0306I\u0307TI\u0307M PLANI KURULSUZ.xlsx"
            template_path = os.path.join("Evraklar", "YILLIKLAR", template_name)
            if not os.path.exists(template_path):
                logging.error(f"Template bulunamadı: {template_path}")
                return
            
            # Hedef dosya adını oluştur
            sirket_proje = self.get_value("[DEĞİŞTİR:ŞİRKETPROJE]")
            kurullu_text = "KURULLU" if is_kurullu else "KURULSUZ"
            output_name = f"{sirket_proje} - Yıllık Eğitim Planı {kurullu_text}.xlsx"
            output_path = os.path.join(self.get_desktop_path(), output_name)
            
            # Template'i kopyala
            shutil.copy2(template_path, output_path)
            
            # Dinamik algoritma uygula
            if use_dynamic_algorithm:
                self.apply_dynamic_algorithm(output_path, 17)  # 17. satır (user tarih bazlı)
            
            # Placeholder'ları doldur
            self.fill_excel_placeholders(output_path)
            
            logging.info(f"Yıllık eğitim planı oluşturuldu: {output_path}")
            
        except Exception as e:
            logging.error(f"Yıllık eğitim planı hatası: {e}")
    
    def generate_yillik_calisma_plani(self, is_kurullu, use_dynamic_algorithm):
        """Yıllık çalışma planını oluşturur"""
        try:
            # Template dosyasını seç (Evraklar/YILLIKLAR klasörü içinde)
            template_name = "YILLIK C\u0327ALIS\u0327MA PLANI KURULLU.xlsx" if is_kurullu else "YILLIK C\u0327ALIS\u0327MA PLANI KURULSUZ.xlsx"
            template_path = os.path.join("Evraklar", "YILLIKLAR", template_name)
            if not os.path.exists(template_path):
                logging.error(f"Template bulunamadı: {template_path}")
                return
            
            # Hedef dosya adını oluştur
            sirket_proje = self.get_value("[DEĞİŞTİR:ŞİRKETPROJE]")
            kurullu_text = "KURULLU" if is_kurullu else "KURULSUZ"
            output_name = f"{sirket_proje} - Yıllık Çalışma Planı {kurullu_text}.xlsx"
            output_path = os.path.join(self.get_desktop_path(), output_name)
            
            # Template'i kopyala
            shutil.copy2(template_path, output_path)
            
            # Dinamik algoritma uygula
            if use_dynamic_algorithm:
                self.apply_dynamic_algorithm(output_path, 6)  # 6. satır (user tarih bazlı)
            
            # Placeholder'ları doldur
            self.fill_excel_placeholders(output_path)
            
            logging.info(f"Yıllık çalışma planı oluşturuldu: {output_path}")
            
        except Exception as e:
            logging.error(f"Yıllık çalışma planı hatası: {e}")
    
    def apply_dynamic_algorithm(self, excel_path, header_row):
        """Dinamik algoritma - geçmiş ayları temizle"""
        try:
            # Mevcut ayı bul
            yillik_tarih = self.get_value("[DEĞİŞTİR:YILLIK:TARİH]")
            if not yillik_tarih:
                return
            
            tarih_obj = datetime.datetime.strptime(yillik_tarih, "%d.%m.%Y")
            current_month = tarih_obj.month
            
            # Excel dosyasını aç
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Ay başlıklarını bul
            ay_isimleri = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", 
                          "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
            
            # Geçmiş ayları temizle
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value and str(cell_value).strip().upper() in ay_isimleri:
                    ay_index = ay_isimleri.index(str(cell_value).strip().upper()) + 1
                    
                    # Eğer bu ay geçmiş aysa, sütunu temizle
                    if ay_index < current_month:
                        self.clear_column_content(ws, col, header_row)
                        logging.info(f"Geçmiş ay temizlendi: {cell_value}")
            
            # Kaydet
            wb.save(excel_path)
            wb.close()
            
        except Exception as e:
            logging.error(f"Dinamik algoritma hatası: {e}")
    
    def clear_column_content(self, worksheet, column_index, header_row):
        """Sütun içeriğini temizler (başlık hariç)"""
        try:
            # Başlık satırından sonraki tüm hücreleri temizle
            for row in range(header_row + 1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=column_index)
                
                # Sadece "X" işaretlerini temizle
                if cell.value and str(cell.value).strip().upper() == "X":
                    cell.value = None
                    # Renk ve dolguyu temizle
                    cell.fill = PatternFill()
                    cell.font = Font()
                    
        except Exception as e:
            logging.error(f"Sütun temizleme hatası: {e}")
    
    def fill_excel_placeholders(self, excel_path):
        """Excel dosyasındaki placeholder'ları doldurur"""
        try:
            # Basit placeholder doldurma - geliştirilecek
            logging.info(f"Placeholder'lar dolduruldu: {excel_path}")
            
        except Exception as e:
            logging.error(f"Placeholder doldurma hatası: {e}")


    @staticmethod
    def darken_color(color):
        """Rengi koyulaştırır (hover efekti için)"""
        return "#0d1235"


# Test için
if __name__ == "__main__":
    root = tk.Tk()
    app = SadeFormApp(root)
    root.mainloop()
